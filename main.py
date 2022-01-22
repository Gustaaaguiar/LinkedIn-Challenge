import random
import qrcode
from openpyxl import load_workbook
import secrets
import os


# declaring the variables globally
item_name = ''
item_description = ''
item_price = 0
item_quantity = 0
item_ID = ''
special_char = '!@#$%&*-+?='
letters = 'abcdefghijklmnopqrstuvwxyz'
numbers = '0123456789'
name_file = 'products.xlsx'
final_quantity = 0


def choices(choice):
    out = 0
    if choice.lower() == 'y':
        out = 1
    elif choice.lower() == 'n':
        out = 2
    else:
        print('Please insert a valid answer')
        choices(choice)

    return out


def loop():
    choice = input('Do you want to register another product? y/n\n')
    # add a loop to the user select if he/she wants to register another product
    x = choices(choice)
    if x == 1:
        product_registration()
    elif x == 2:
        exit()


def register_user():
    username = str(input('Insert the username: '))  # register the username
    regis = int(input('You want to:\n1-Use your password \n2-Use an auto generated password '))
    password = ''

    if regis == 1:
        password = str(input('Insert the password: '))
        print(f'Your password is: {password}')  # register the password

    elif regis == 2:
        alphabet = letters + letters.upper() + special_char + numbers
        password = password.join(secrets.choice(alphabet) for _ in range(10))  # generate password with the number of
        # digits I want
        print(f'Your password is: {password}')  # register the password

    else:
        print('please select a valid answer')
        register_user()

    print(f'The username is: {username}\nThe password is: {password}')


def remove(sheet, row):
    # iterate the row object
    for cell in row:
        # check the value of each cell in
        # the row, if any of the value is not
        # None return without removing the row
        if cell.value != None:
            return
    # get the row number from the first cell
    # and remove the row
    sheet.delete_rows(row[0].row, 1)


def checking():  # check if there's any empty cell and delete it
    # if os.getcwd() != 'E:\PythonProjs\LinkerdIn-challenge\output':
    #     os.chdir('output')

    workbook = load_workbook(filename=name_file, data_only=True)
    sheet = workbook.active

    for row in sheet:
        remove(sheet, row)

    workbook.save(filename=name_file)


def set_name():
    global item_name
    item_name = input('Type the product name: ')  # add a function to not allow the name to be null
    item_name.lower()
    if item_name == '':
        print('Name cannot be null')
        set_name()

    workbook = load_workbook(filename=name_file, data_only=True)
    sheet = workbook.active

    for items in sheet.iter_cols(min_col=2, max_col=2, values_only=True):
        for products in items:
            if products == item_name:
                print('Product already registered')
                loop()


def set_price():
    global item_price
    try:  # error handling
        item_price = float(input('Type the price of the product: ').replace(',', '.'))
    except ValueError:  # error when you try to insert a non number inside a float or int
        print('The value inserted was does not correspond to a number')
    item_price = "{:.2f}".format(item_price)  # force to print the number with 2 decimal places
    if item_price == '':
        print('Price cannot be null')
        set_price()


def set_quantity():
    global item_quantity
    try:  # error handling
        item_quantity = int(input('Type the quantity available: '))
    except ValueError:  # error when you try to insert a non number inside a float or int
        print('The value inserted was does not correspond to an integer')
        set_quantity()
        if item_quantity == '':
            print('Quantity cannot be null')
            set_quantity()


def set_ID():
    global item_ID
    ID = ''

    for i in range(5):
        number = random.randint(0, 9)  # create a random number between 0 and 9 to create the ID
        ID += str(number)  # get the 'number' variable and append it to the 'ID' variable
    item_ID = f'#{ID}'  # just put a '#' before the ID

    workbook = load_workbook(filename=name_file, data_only=True)
    sheet = workbook.active

    for items in sheet.iter_cols(min_col=5, max_col=5, values_only=True):  # check if the ID is available and change
        # if needed
        for products in items:
            if item_ID == products:
                set_ID()


def create_qr():
    qr = qrcode.QRCode(
        version=1,  # determines the size of the qrcode
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=4,  # determines the size of the boxes
        border=2,  # determines the thickness of the border
    )

    qr.add_data(f'Item name: {item_name}\nItem description: {item_description}\nItem price: R${item_price}\n'
                f'Item quantity: {item_quantity}\nItem ID: {item_ID}')  # add the content to the qrcode
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white").convert('RGB')

    img.save(f"qrcode/ {item_ID}.png")  # saves the image as and .png file on the folder qrcodes


def add_to_spreadsheet():
    workbook = load_workbook(filename=name_file)
    sheet = workbook.active  # open the spreadsheet and detects in what page it is

    column = sheet['B']
    last_column = len(column) + 1  # check the last completed cell and start appending data on the next

    sheet[f"B{last_column}"] = item_name.lower()  # always append the name with lower case to make checking easier
    sheet[f"C{last_column}"] = item_price
    sheet[f"D{last_column}"] = item_quantity
    sheet[f"E{last_column}"] = item_ID

    workbook.save(filename=name_file)  # save the spreadsheet


def product_registration():  # just run all the important functions
    global item_name, item_description, item_price, item_quantity, item_ID

    checking()
    set_name()
    item_description = input('Type a description for the product (optional): ')
    set_price()
    set_quantity()
    set_ID()
    add_to_spreadsheet()
    create_qr()
    print(f'Item name: {item_name}\nItem description: {item_description}\nItem price: {item_price}\nItem quantity: '
          f'{item_quantity}\nItem ID: {item_ID}')  # print the data inserted to ease the view of it
    loop()


def checking_sell():  # wip
    workbook = load_workbook(filename=name_file)
    sheet = workbook.active  # open the spreadsheet and detects in what page it is

    row = 1

    # inputting the code
    product = str(input('Type the code of the product: '))
    product = f'#{product}'

    # checking if the code is valid
    if len(product) != 6:
        print('The code inserted is invalid!')
        checking_sell()

    else:
        for items in sheet.iter_cols(min_col=5, max_col=5, values_only=True):  # checking the ID column
            for ids in items:
                if ids == product:
                    sell(row)
                    break
                else:
                    row += 1
                    if sheet.max_row != row:
                        pass
                    else:
                        print('The code does not correspond to a registered product!')
                        choice = input('Do you want to register a new product? y/n')

                        x = choices(choice)

                        if x == 1:
                            product_registration()
                        elif x == 2:
                            choice = input('Do you want to sell another product? y/n')

                            y = choices(choice)

                            if y == 1:
                                checking_sell()
                            elif y == 2:
                                exit()
                            else:
                                print('Please type a valid answer!')
                                choices(choice)
                        else:
                            print('Please type a valid answer!')
                            choices(choice)


def sell(sel_row):
    global final_quantity

    workbook = load_workbook(filename=name_file)
    sheet = workbook.active  # open the spreadsheet and detects in what page it is

    price = 0

    out = []

    for c in range(1, sheet.max_column + 1):
        out.append(sheet.cell(sel_row, c).value)

    out.pop(0)

    if out[2] <= 0:
        print('Product not available in our stock')
        choice = input('Do you want to sell another product? y/n\n')

        x = choices(choice)

        if x == 1:
            checking_sell()
        elif x == 2:
            exit()
        else:
            print('Please type a valid answer!')
            choices(choice)
    else:
        qt_sold = int(input('Type the quantity to be sold: '))

        if qt_sold > out[2]:
            print('The quantity to be sold is not available in our stock!')

            choice = input('Do you want to sell another product? y/n\n')

            x = choices(choice)

            if x == 1:
                checking_sell()
            elif x == 2:
                exit()
            else:
                print('Please type a valid answer!')
                choices(choice)
        else:
            item_pr = float(out[1])
            price += qt_sold*item_pr
            price = "{:.2f}".format(price)

            final_quantity = qt_sold-out[2]
            sheet[f'D{sel_row}'] = final_quantity

            workbook.save(filename=name_file)

            choice = input('Do you want to sell another product? y/n\n')

            x = choices(choice)

            if x == 1:
                checking_sell()
            elif x == 2:
                price = str(price)
                price = price.replace('.', ',')
                print(f'The final price is R${price}')
                input('Press enter to exit.')
                exit()
            else:
                print('Please type a valid answer!')
                choices(choice)


os.chdir('output')
checking_sell()
