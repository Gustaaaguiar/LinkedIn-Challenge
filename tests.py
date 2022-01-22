# read qrcode

from openpyxl import load_workbook
import os


os.chdir('output')

workbook = load_workbook(filename='products.xlsx')
sheet = workbook.active

for items in sheet.iter_cols(min_col=5, max_col=5, values_only=True):
    for ids in items:
        if ids == 2:
            row = ids


out = []

for c in range(5):
    out.append(sheet.cell(3, c+1).value)

out.pop(0)

# print(out)

price, quantity = float(out[1]), out[2]

print(type(price))

